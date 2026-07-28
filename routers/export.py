import csv
import io
import os
import shutil
import tempfile
from datetime import date, datetime
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from database import (
    DB_PATH, get_db, recreate_engine, Registro, TipoCirugia,
)
from auth import get_current_user, get_admin_user, Usuario

router = APIRouter(prefix="/api/export", tags=["export"])

# ── Explicit column order ────────────────────────────────────────────────────
COMMON_COLS = [
    'id', 'nhc', 'fecha_intervencion', 'fecha_nacimiento', 'edad', 'sexo',
    'asa', 'cirujano', 'ingreso_cma', 'diagnostico', 'intervencion',
    'urgencia', 'abordaje', 'conversion', 'tiempo_quirurgico',
    'clavien_dindo', 'tipo_complicacion', 'intervencionismo',
    'reintervencion', 'tipo_reintervencion', 'mortalidad', 'causa_mortalidad',
    'fecha_alta', 'estancia', 'reingreso_30d', 'observaciones',
    'created_by', 'created_at',
]

COLORRECTAL_COLS = [
    'id', 'nhc', 'fecha_intervencion', 'fecha_nacimiento', 'edad', 'sexo',
    'asa', 'cirujano', 'ingreso_cma', 'diagnostico', 'intervencion',
    'urgencia', 'abordaje', 'conversion', 'tiempo_quirurgico',
    'estoma_proteccion',
    'clavien_dindo', 'dehiscencia', 'tipo_dehiscencia',
    'tipo_complicacion', 'intervencionismo',
    'reintervencion', 'tipo_reintervencion', 'mortalidad', 'causa_mortalidad',
    'fecha_alta', 'estancia', 'reingreso_30d',
    # Oncológico
    't_tnm', 'n_tnm', 'm_tnm', 'estadio_tnm',
    'localizacion', 'distancia_margen_anal',
    'neoadyuvancia', 'tipo_neoadyuvancia', 'pcr',
    'tipo_histologico', 'grado', 'margenes_libres',
    'ganglios_analizados', 'ganglios_positivos',
    'invasion_linfovascular', 'invasion_perineural', 'msi',
    'adyuvancia', 'tipo_adyuvancia',
    # Seguimiento
    'recidiva_3m', 'tipo_recidiva_3m',
    'recidiva_6m', 'tipo_recidiva_6m',
    'recidiva_12m', 'tipo_recidiva_12m',
    'recidiva_18m', 'tipo_recidiva_18m',
    'recidiva_24m', 'tipo_recidiva_24m',
    'recidiva_36m', 'tipo_recidiva_36m',
    'recidiva_48m', 'tipo_recidiva_48m',
    'recidiva_60m', 'tipo_recidiva_60m',
    'fecha_exitus',
    'observaciones', 'created_by', 'created_at',
]

def _columnas(tipo: TipoCirugia) -> list[str]:
    """Los tipos con seguimiento oncológico llevan además el bloque TNM."""
    return COLORRECTAL_COLS if tipo.tiene_oncologico else COMMON_COLS


def _tipos_a_exportar(db, tabla: str) -> list[TipoCirugia]:
    """`tabla` es "todas" o el slug de un tipo."""
    q = db.query(TipoCirugia).order_by(TipoCirugia.orden, TipoCirugia.id)
    if tabla and tabla != "todas":
        q = q.filter(TipoCirugia.slug == tabla)
    return q.all()


def _get_rows(db, tipo_id, fecha_desde, fecha_hasta):
    q = db.query(Registro).filter(Registro.tipo_id == tipo_id)
    if fecha_desde:
        q = q.filter(Registro.fecha_intervencion >= fecha_desde)
    if fecha_hasta:
        q = q.filter(Registro.fecha_intervencion <= fecha_hasta)
    return q.order_by(Registro.fecha_intervencion).all()


def _row_to_dict(obj, cols):
    result = {}
    for col in cols:
        val = getattr(obj, col, None)
        if isinstance(val, date):
            val = val.isoformat()
        result[col] = val if val is not None else ''
    return result


# ── CSV ──────────────────────────────────────────────────────────────────────
@router.get("/csv")
def export_csv(
    tabla: str = Query("todas"),
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user),
):
    output = io.StringIO()
    tipos = _tipos_a_exportar(db, tabla)
    todas = tabla == "todas"

    writer = None
    for tipo in tipos:
        # Al exportar todo se usan las columnas comunes, para que salga un CSV único
        write_cols = COMMON_COLS if todas else _columnas(tipo)
        rows = _get_rows(db, tipo.id, fecha_desde, fecha_hasta)
        if not rows:
            continue
        if writer is None:
            fieldnames = (['tipo_cirugia'] + write_cols) if todas else write_cols
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
        for r in rows:
            d = _row_to_dict(r, write_cols)
            if todas:
                d = {'tipo_cirugia': tipo.nombre, **d}
            writer.writerow({k: d.get(k, '') for k in writer.fieldnames})

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=coloproctologia_{tabla}.csv"},
    )


# ── XLSX ─────────────────────────────────────────────────────────────────────
@router.get("/xlsx")
def export_xlsx(
    tabla: str = Query("todas"),
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    for tipo in _tipos_a_exportar(db, tabla):
        cols = _columnas(tipo)
        # El color de cabecera es el que el admin haya elegido para el tipo
        color_hdr = "FF" + (tipo.color or "#0D2B4E").lstrip("#").upper()
        # Excel limita los nombres de hoja a 31 caracteres y prohíbe : \ / ? * [ ]
        titulo = tipo.nombre[:31]
        for c in ':\\/?*[]':
            titulo = titulo.replace(c, '-')
        ws = wb.create_sheet(title=titulo)
        rows = _get_rows(db, tipo.id, fecha_desde, fecha_hasta)
        if not rows:
            ws.append(["Sin datos para el período seleccionado"])
            continue

        dicts = [_row_to_dict(r, cols) for r in rows]

        # Header row
        ws.append(cols)
        hdr_fill = PatternFill("solid", fgColor=color_hdr)
        hdr_font = Font(bold=True, color="FFFFFFFF", size=10)
        for ci, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data rows
        LIGHT_GRAY = PatternFill("solid", fgColor="FFF5F5F5")
        for ri, d in enumerate(dicts, 2):
            row_fill = LIGHT_GRAY if ri % 2 == 0 else None
            for ci, col_name in enumerate(cols, 1):
                cell = ws.cell(row=ri, column=ci, value=d.get(col_name, ''))
                if row_fill:
                    cell.fill = row_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Column widths — wider for date/text cols
        for ci, col_name in enumerate(cols, 1):
            if 'fecha' in col_name or 'observacion' in col_name or col_name in ('diagnostico','intervencion','tipo_reintervencion','causa_mortalidad'):
                ws.column_dimensions[get_column_letter(ci)].width = 22
            elif col_name in ('id','asa','sexo','t_tnm','n_tnm','m_tnm','grado'):
                ws.column_dimensions[get_column_letter(ci)].width = 8
            else:
                ws.column_dimensions[get_column_letter(ci)].width = 16
        ws.row_dimensions[1].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=coloproctologia_{tabla}.xlsx"},
    )


# ── BACKUP ───────────────────────────────────────────────────────────────────
@router.get("/backup")
def download_backup(_: Usuario = Depends(get_admin_user)):
    """Stream the raw SQLite database file. Admin only."""
    if not os.path.exists(DB_PATH):
        raise HTTPException(status_code=404, detail="Base de datos no encontrada")
    filename = f"backup_coloproctologia_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.db"
    return FileResponse(
        path=DB_PATH,
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── RESTORE ──────────────────────────────────────────────────────────────────
SQLITE_MAGIC = b"SQLite format 3\x00"

@router.post("/restore")
async def restore_backup(
    file: UploadFile = File(...),
    _: Usuario = Depends(get_admin_user),
):
    """Replace the current database with an uploaded .db backup. Admin only."""
    content = await file.read()
    if not content.startswith(SQLITE_MAGIC):
        raise HTTPException(
            status_code=400,
            detail="El archivo no es una base de datos SQLite válida",
        )

    # Write to a temp file first, then atomically replace
    db_dir = os.path.dirname(DB_PATH)
    os.makedirs(db_dir, exist_ok=True)
    tmp_fd, tmp_path = tempfile.mkstemp(dir=db_dir, suffix=".db.tmp")
    try:
        with os.fdopen(tmp_fd, "wb") as f:
            f.write(content)
        # Close all SQLAlchemy connections before replacing the file
        recreate_engine()
        shutil.move(tmp_path, DB_PATH)
        # Re-connect to the restored file
        recreate_engine()
    except Exception as e:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise HTTPException(status_code=500, detail=f"Error al restaurar: {e}")

    return {"ok": True, "message": "Base de datos restaurada correctamente"}


# ── IMPORT CSV ───────────────────────────────────────────────────────────────
# Campos que no se copian: los asigna la base de datos
_SKIP = {"id", "created_at", "tipo_id"}

DATE_COLS = {
    "fecha_intervencion", "fecha_nacimiento", "fecha_alta", "fecha_exitus",
}


def _parse_date(val: str):
    if not val or val.strip() == "":
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(val.strip(), fmt).date()
        except ValueError:
            pass
    return None


def _buscar_tipo(db: Session, valor: str):
    """Acepta el slug o el nombre visible del tipo, sin distinguir mayúsculas."""
    if not valor:
        return None
    v = valor.strip().lower()
    for t in db.query(TipoCirugia).all():
        if t.slug.lower() == v or t.nombre.lower() == v:
            return t
    return None


@router.post("/import-csv")
async def import_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_admin_user),
):
    """Add records from a CSV export without deleting existing rows. Admin only."""
    raw = await file.read()
    try:
        text = raw.decode("utf-8-sig")  # handle BOM from Excel CSV exports
    except UnicodeDecodeError:
        text = raw.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    columnas_validas = {c.key for c in Registro.__table__.columns}

    inserted = 0
    skipped = 0
    errors = []

    for i, row in enumerate(reader, start=2):   # 2 = primera fila de datos
        try:
            tipo = _buscar_tipo(db, row.get("tipo_cirugia", ""))
            if not tipo:
                errors.append(
                    f"Fila {i}: tipo de cirugía desconocido "
                    f"('{row.get('tipo_cirugia', '')}')"
                )
                skipped += 1
                continue

            kwargs: dict = {}
            for col, val in row.items():
                col = (col or "").strip()
                if col in _SKIP or col == "tipo_cirugia" or col not in columnas_validas:
                    continue
                if val == "" or val is None:
                    kwargs[col] = None
                elif col in DATE_COLS:
                    kwargs[col] = _parse_date(val)
                else:
                    kwargs[col] = val

            kwargs["tipo_id"] = tipo.id
            kwargs["created_by"] = current_user.username
            kwargs["created_at"] = datetime.utcnow()
            db.add(Registro(**kwargs))
            inserted += 1
        except Exception as e:
            errors.append(f"Fila {i}: {e}")
            skipped += 1

    db.commit()

    result = {"ok": True, "insertados": inserted, "omitidos": skipped}
    if errors:
        result["errores"] = errors[:20]  # cap at 20 to avoid huge response
    return result
