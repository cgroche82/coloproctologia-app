"""
Importación del Excel de intervenidos que genera la aplicación de Lista de Espera.

Funciona en dos pasos deliberadamente:

    1. /analizar   lee el fichero y devuelve una propuesta. NO escribe nada.
    2. /confirmar  recibe lo que el usuario haya revisado y lo graba.

El motivo es que ese Excel es de planificación, no un parte quirúrgico: los
diagnósticos van en texto libre y los procedimientos a veces expresan una duda
sin resolver ("TAMIS VS RESECCION ANTERIOR"). Convertirlo automáticamente
metería datos clínicos incorrectos sin que nadie se entere, así que todo lo
dudoso se propone y lo confirma una persona.
"""

import difflib
import io
import re
import unicodedata
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session

from database import get_db, Registro, TipoCirugia, Diagnostico, Intervencion, Cirujano
from auth import get_admin_user, Usuario

router = APIRouter(prefix="/api/import", tags=["importador"])

# Confirmado con el usuario: en el Excel de Lista de Espera H=Hombre, M=Mujer.
# Interpretarlo al revés invertiría el sexo de decenas de pacientes en silencio.
SEXO = {"H": "Masculino", "M": "Femenino"}

# Sólo se sugiere una equivalencia de diagnóstico por encima de este parecido.
# Con 0.75 llegaba a proponer "Neoplasia de Recto" para "NEOPLASIA DERECHA".
UMBRAL_SUGERENCIA = 0.90

# Palabras clave para adivinar el grupo de las filas sin clasificar
PISTAS_GRUPO = [
    ("neoplasias", ("NEOPLASIA", "NOPLASIA", "NEOPLSASIA", "TUMOR", "CANCER",
                    "CARCINOMA", "ADENOCARCINOMA", "RECIDIVA", "POLIPO",
                    "TUMORACION")),
    ("eii", ("COLITIS ULCEROSA", "CROHN", "INFLAMATORIA")),
    ("proctologia", ("PROLAPSO", "FISTULA", "HEMORROID", "FISURA", "SINUS",
                     "ABSCESO", "CONDILOMA", "FOURNIER")),
    ("neuromodulacion", ("INCONTINENCIA", "RECTOCELE")),
    ("reconstrucciones", ("CIERRE", "RECONSTRUC", "HARTMANN", "ESTOMA",
                          "ILEOSTOMIA", "COLOSTOMIA")),
    ("colon_benigno", ("DIVERTICUL", "ESTENOSIS", "OGILVIE", "ISQUEMICA",
                       "OBSTRUCCION")),
]

# Abreviaturas que aparecen en los datos reales
_ABREVIATURAS = [
    (r"^N\s+", "NEOPLASIA "),          # "N COLON DERECHO"
    (r"^NEO\s+", "NEOPLASIA "),        # "NEO COLON DERECHO"
    (r"\bIZDO\b", "IZQUIERDO"),
    (r"\bIZDA\b", "IZQUIERDA"),
    (r"\bDCHO\b", "DERECHO"),
    (r"\bDCHA\b", "DERECHA"),
]


def normalizar(texto) -> str:
    """Mayúsculas, sin tildes y sin espacios de más."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode()
    return " ".join(t.upper().split())


def _expandir(texto: str) -> str:
    for patron, reemplazo in _ABREVIATURAS:
        texto = re.sub(patron, reemplazo, texto)
    return texto


def _parsear_fecha(valor):
    if valor is None or str(valor).strip() == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if hasattr(valor, "year") and hasattr(valor, "month"):
        return valor
    s = str(valor).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _partir_alternativas(procedimiento: str) -> list[str]:
    """
    "TAMIS VS RESECCION ANTERIOR" -> ["TAMIS", "RESECCION ANTERIOR"].

    Un procedimiento con VS es lo que se pensaba hacer, no lo que se hizo, así
    que se ofrecen las opciones para que el usuario marque la correcta.
    """
    if not procedimiento:
        return []
    partes = re.split(r"\s+VS\.?\s+", procedimiento.strip(), flags=re.IGNORECASE)
    return [p.strip() for p in partes if p.strip()] if len(partes) > 1 else []


# ── Esquemas de la respuesta ─────────────────────────────────────────────────
class FilaConfirmada(BaseModel):
    nhc: str
    fecha_intervencion: str
    tipo_id: int
    diagnostico: Optional[str] = None
    intervencion: Optional[str] = None
    cirujano: Optional[str] = None
    edad: Optional[int] = None
    sexo: Optional[str] = None
    observaciones: Optional[str] = None


class Confirmacion(BaseModel):
    filas: list[FilaConfirmada]


# ── Paso 1: analizar ─────────────────────────────────────────────────────────
@router.post("/lista-espera/analizar")
async def analizar(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: Usuario = Depends(get_admin_user),
):
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(await file.read()), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el Excel: {e}")

    tipos = db.query(TipoCirugia).filter(TipoCirugia.activo == True).all()
    tipos_por_nombre = {normalizar(t.nombre): t for t in tipos}
    tipos_por_slug = {t.slug: t for t in tipos}

    diags = db.query(Diagnostico).filter(Diagnostico.activo == True).all()
    diag_por_norm = {}
    for d in diags:
        diag_por_norm.setdefault(normalizar(d.nombre), d)

    cirujanos = db.query(Cirujano).filter(Cirujano.activo == True).all()
    # "GRACIA" debe casar con "DR. GRACIA": se compara sin el tratamiento.
    # Ojo con el orden de la alternancia: (DR|DRA) haría que "DRA. GASCÓN" sólo
    # perdiera el "DR" y quedara como "A. GASCON", que no casa con nada.
    cir_por_apellido = {
        re.sub(r"^DRA?\.?\s*", "", normalizar(c.nombre)): c.nombre
        for c in cirujanos
    }

    # Episodios ya guardados, para no duplicar
    existentes = {
        (r.nhc.strip(), r.fecha_intervencion.isoformat())
        for r in db.query(Registro).all()
        if r.nhc and r.fecha_intervencion
    }

    filas = []
    vistos_en_fichero: dict[tuple, int] = {}
    cirujanos_desconocidos: set[str] = set()

    for hoja in wb.sheetnames:
        ws = wb[hoja]
        cabecera = [str(c.value).strip() if c.value else "" for c in ws[1]]
        col = {n: i for i, n in enumerate(cabecera) if n}

        def val(fila, nombre):
            i = col.get(nombre)
            return fila[i] if i is not None and i < len(fila) else None

        tipo_hoja = tipos_por_nombre.get(normalizar(hoja))

        for n_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(v not in (None, "") for v in fila):
                continue

            nhc = str(val(fila, "NHC") or "").strip()
            fecha = _parsear_fecha(val(fila, "F. Programación"))
            diag_bruto = str(val(fila, "Diagnóstico") or "").strip()
            proc_bruto = str(val(fila, "Procedimiento") or "").strip()

            aviso = []
            if not nhc:
                aviso.append("Sin NHC")
            if not fecha:
                aviso.append("Sin fecha de intervención")

            # ── Grupo ──
            tipo = tipo_hoja
            grupo_sugerido = False
            if not tipo:
                d = _expandir(normalizar(diag_bruto))
                for slug, palabras in PISTAS_GRUPO:
                    if any(p in d for p in palabras) and slug in tipos_por_slug:
                        tipo = tipos_por_slug[slug]
                        grupo_sugerido = True
                        break

            # ── Diagnóstico ──
            diag_sugerido, confianza = None, None
            if diag_bruto:
                d_norm = _expandir(normalizar(diag_bruto))
                if d_norm in diag_por_norm:
                    diag_sugerido, confianza = diag_por_norm[d_norm].nombre, 1.0
                else:
                    # Sólo se compara con los diagnósticos del grupo propuesto
                    candidatos = [normalizar(x.nombre) for x in diags
                                  if not tipo or x.tipo_id == tipo.id]
                    cerca = difflib.get_close_matches(
                        d_norm, candidatos, n=1, cutoff=UMBRAL_SUGERENCIA)
                    if cerca:
                        diag_sugerido = diag_por_norm[cerca[0]].nombre
                        confianza = round(
                            difflib.SequenceMatcher(None, d_norm, cerca[0]).ratio(), 2)

            # ── Intervención ──
            alternativas = _partir_alternativas(proc_bruto)
            interv_sugerida = None
            if proc_bruto and not alternativas:
                p_norm = normalizar(proc_bruto)
                opciones = {normalizar(i.nombre): i.nombre for i in
                            db.query(Intervencion).filter(Intervencion.activo == True).all()}
                if p_norm in opciones:
                    interv_sugerida = opciones[p_norm]
                else:
                    cerca = difflib.get_close_matches(
                        p_norm, list(opciones), n=1, cutoff=UMBRAL_SUGERENCIA)
                    if cerca:
                        interv_sugerida = opciones[cerca[0]]

            # ── Cirujano ──
            cir_bruto = str(val(fila, "Cirujano") or "").strip()
            cirujano = cir_por_apellido.get(normalizar(cir_bruto)) if cir_bruto else None
            if cir_bruto and not cirujano:
                cirujanos_desconocidos.add(cir_bruto)

            # ── Duplicados ──
            clave = (nhc, fecha.isoformat()) if nhc and fecha else None
            estado = "nuevo"
            if clave:
                if clave in existentes:
                    estado = "ya_en_base"
                elif clave in vistos_en_fichero:
                    estado = "repetido_en_excel"
                    aviso.append(f"Mismo NHC y fecha que la fila {vistos_en_fichero[clave]}")
                else:
                    vistos_en_fichero[clave] = n_fila

            edad = val(fila, "Edad")
            filas.append({
                "fila": n_fila,
                "hoja": hoja,
                "nhc": nhc,
                "fecha_intervencion": fecha.isoformat() if fecha else None,
                "edad": int(edad) if isinstance(edad, (int, float)) else None,
                "sexo": SEXO.get(normalizar(val(fila, "Sexo"))),
                "cirujano": cirujano,
                "cirujano_bruto": cir_bruto or None,
                "diagnostico_bruto": diag_bruto or None,
                "diagnostico_sugerido": diag_sugerido,
                "confianza": confianza,
                "procedimiento_bruto": proc_bruto or None,
                "intervencion_sugerida": interv_sugerida,
                "alternativas": alternativas,
                "observaciones": str(val(fila, "Observaciones") or "").strip() or None,
                "tipo_id": tipo.id if tipo else None,
                "tipo_nombre": tipo.nombre if tipo else None,
                "grupo_sugerido": grupo_sugerido,
                "estado": estado,
                "importable": bool(nhc and fecha and tipo),
                "avisos": aviso,
            })

    return {
        "filas": filas,
        "resumen": {
            "total": len(filas),
            "nuevos": sum(1 for f in filas if f["estado"] == "nuevo" and f["importable"]),
            "ya_en_base": sum(1 for f in filas if f["estado"] == "ya_en_base"),
            "repetidos_en_excel": sum(1 for f in filas if f["estado"] == "repetido_en_excel"),
            "no_importables": sum(1 for f in filas if not f["importable"]),
            "grupo_sugerido": sum(1 for f in filas if f["grupo_sugerido"]),
            "sin_diagnostico": sum(1 for f in filas
                                   if f["diagnostico_bruto"] and not f["diagnostico_sugerido"]),
            "con_alternativas": sum(1 for f in filas if f["alternativas"]),
        },
        "cirujanos_desconocidos": sorted(cirujanos_desconocidos),
    }


# ── Paso 2: confirmar ────────────────────────────────────────────────────────
@router.post("/lista-espera/confirmar")
def confirmar(
    datos: Confirmacion,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_admin_user),
):
    """Graba lo que el usuario haya revisado. La columna de nombres del Excel
    no llega hasta aquí: la aplicación sólo guarda NHC."""
    existentes = {
        (r.nhc.strip(), r.fecha_intervencion.isoformat())
        for r in db.query(Registro).all()
        if r.nhc and r.fecha_intervencion
    }

    insertados, omitidos, errores = 0, 0, []
    for i, f in enumerate(datos.filas, start=1):
        try:
            fecha = _parsear_fecha(f.fecha_intervencion)
            if not f.nhc or not fecha:
                omitidos += 1
                continue
            clave = (f.nhc.strip(), fecha.isoformat())
            if clave in existentes:      # protege de un doble envío
                omitidos += 1
                continue
            if not db.query(TipoCirugia).filter(TipoCirugia.id == f.tipo_id).first():
                errores.append(f"Fila {i}: grupo inexistente")
                omitidos += 1
                continue

            db.add(Registro(
                tipo_id=f.tipo_id, nhc=f.nhc.strip(), fecha_intervencion=fecha,
                edad=f.edad, sexo=f.sexo, cirujano=f.cirujano,
                diagnostico=f.diagnostico, intervencion=f.intervencion,
                observaciones=f.observaciones,
                created_by=current_user.username, created_at=datetime.utcnow(),
            ))
            existentes.add(clave)
            insertados += 1
        except Exception as e:            # noqa: BLE001 - se informa fila a fila
            errores.append(f"Fila {i}: {e}")
            omitidos += 1

    db.commit()
    return {"ok": True, "insertados": insertados, "omitidos": omitidos,
            "errores": errores[:20]}
